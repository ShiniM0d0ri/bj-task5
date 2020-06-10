import test_info
import openpyxl

wb = openpyxl.load_workbook('Distributed Data_Client Master List.xlsx')

print("Enter starting row and ending row :D ")
rowinit=int(input())
rowend=int(input())


ws=wb.active
for i in range(rowinit,rowend+1):
    row=str(i)
    comp_name=ws['D'+row].value
    
    print("\tCompany Name: ",comp_name)
    cat=test_info.spider(comp_name)
    ws['K'+row]=cat
    print("\tEntered "+cat+" Successfully......")

wb.save("updated.xlsx")
print("updated.xlsx saved successfully")
print("\t\t:) :) :)")